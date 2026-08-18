from io import BytesIO

from openpyxl import load_workbook
import pytest

from src.m6_workout_plan import generate_weekly_plan
from src.program_schedule import program_window
from src.workout_plan_excel_export import export_sessions_workbook, export_sessions_workbook_bytes


VALIDATED_PLAN = {
    "plan_json": {
        "sessions": [
            {
                "scheduled_date": "2026-08-14",
                "day": "Day 1",
                "focus": "Lower-body strength",
                "exercises": ["Goblet squat", "Dumbbell Romanian deadlift"],
                "sets_reps": "3 sets of 8 reps",
                "adjustment": "Reduce working sets by 20%.",
            },
            {
                "scheduled_date": "2026-08-16",
                "day": "Day 2",
                "focus": "Badminton conditioning",
                "exercises": ["Badminton footwork intervals"],
                "sets_reps": "20 minutes",
                "adjustment": "Train as planned.",
            },
        ]
    }
}


def test_export_sessions_workbook_writes_one_readable_row_per_session(tmp_path):
    output_path = export_sessions_workbook(
        VALIDATED_PLAN,
        tmp_path / "yu_latest_workout_plan.xlsx",
    )

    workbook = load_workbook(output_path)
    sheet = workbook["Sessions"]

    assert workbook.sheetnames == ["Sessions"]
    assert [cell.value for cell in sheet[1]] == [
        "Date",
        "Day",
        "Focus",
        "Exercises",
        "Sets/Reps",
        "Adjustment",
    ]
    assert [cell.value for cell in sheet[2]] == [
        "2026-08-14",
        "Day 1",
        "Lower-body strength",
        "Goblet squat\nDumbbell Romanian deadlift",
        "3 sets of 8 reps",
        "Reduce working sets by 20%.",
    ]
    assert sheet["D2"].alignment.wrap_text is True
    assert sheet.freeze_panes == "A2"


def test_export_sessions_workbook_rejects_an_empty_session_list(tmp_path):
    with pytest.raises(ValueError, match="no sessions"):
        export_sessions_workbook(
            {"plan_json": {"sessions": []}},
            tmp_path / "empty_plan.xlsx",
        )


def test_export_sessions_workbook_bytes_is_a_downloadable_workbook():
    workbook_bytes = export_sessions_workbook_bytes(VALIDATED_PLAN)

    workbook = load_workbook(BytesIO(workbook_bytes))

    assert workbook["Sessions"]["A2"].value == "2026-08-14"
    assert workbook["Sessions"]["C2"].value == "Lower-body strength"


def test_exported_workbook_keeps_a_friday_inquiry_as_day_one():
    program = program_window("2026-08-14", duration_weeks=8)
    plan = generate_weekly_plan(
        profile={"weekly_availability": "3 days/week"},
        goal={
            "id": "goal-1",
            "goal_type": "sport_conditioning",
            "plan_duration_weeks": 8,
            "goal_details": {"athlete_type": "badminton"},
        },
        readiness={"readiness_score": 85, "band": "train_as_planned", "safety_triggered": False},
        week_start=program["program_start_date"],
    )

    workbook = load_workbook(BytesIO(export_sessions_workbook_bytes({"plan_json": plan})))
    sheet = workbook["Sessions"]

    assert [cell.value for cell in sheet[1][:2]] == ["Date", "Day"]
    assert [cell.value for cell in sheet[2][:2]] == ["2026-08-14", "Day 1"]
