from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


SESSION_HEADERS = ["Day", "Focus", "Exercises", "Sets/Reps", "Adjustment"]
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def export_sessions_workbook(plan: dict[str, Any], output_path: Path) -> Path:
    """Write a sessions-only workout-plan workbook and return its saved path."""
    workbook = _build_sessions_workbook(plan)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def export_sessions_workbook_bytes(plan: dict[str, Any]) -> bytes:
    """Return a sessions-only workbook for an in-browser download."""
    output = BytesIO()
    _build_sessions_workbook(plan).save(output)
    return output.getvalue()


def _build_sessions_workbook(plan: dict[str, Any]) -> Workbook:
    sessions = (plan.get("plan_json") or {}).get("sessions") or []
    if not sessions:
        raise ValueError("Workout plan has no sessions to export.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sessions"
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.append(SESSION_HEADERS)

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for session in sessions:
        exercises = "\n".join(str(exercise) for exercise in session.get("exercises") or [])
        sheet.append(
            [
                session.get("day", ""),
                session.get("focus", ""),
                exercises,
                session.get("sets_reps", ""),
                session.get("adjustment", ""),
            ]
        )

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row_number in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_number].height = 34

    for column, width in {
        "A": 14,
        "B": 28,
        "C": 42,
        "D": 24,
        "E": 38,
    }.items():
        sheet.column_dimensions[column].width = width
    return workbook
